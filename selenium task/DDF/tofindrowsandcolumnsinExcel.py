import openpyxl

excel_path="C:\\Users\\pc\\Desktop\\DDF.xlsx"
work_book=openpyxl.load_workbook(excel_path)
sheet=work_book["Sheet1"]
row_size=sheet.max_row
print("Row size is ",row_size)
col_size=sheet.max_column
print("Column size is " ,col_size)
