import openpyxl


def row_Count(excel_Path,sheet):
    wb=openpyxl.load_workbook(excel_Path)
    sh=wb[sheet]
    rows=sh.max_row
    return rows

def col_Count(excel_Path,sheet):
    wb=openpyxl.load_workbook(excel_Path)
    sh=wb[sheet]
    cols=sh.max_column
    return cols

def read_Excel(excel_Path,sheet,r,c):
    wb=openpyxl.load_workbook(excel_Path)
    sh=wb[sheet]
    data=sh.cell(row=r,column=c).value
    return data

def write_Excel(excel_Path,sheet,r,c,val):
    wb=openpyxl.load_workbook(excel_Path)
    sh=wb[sheet]
    sh.cell(row=r,column=c).value=val
    wb.save(excel_Path)