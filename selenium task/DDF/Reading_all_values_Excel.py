import openpyxl

excel_Path='C:\\Users\\pc\\Desktop\\DDF.xlsx'

wb=openpyxl.load_workbook(excel_Path)
sh=wb.active
row_size=sh.max_row
col_size=sh.max_column
print(sh.cell(row=2,column=1).value)
print('=========')
for r in range(2,row_size+1):
    for c in range(1,col_size+1):
        print(sh.cell(row=r,column=c).value, end=" ")
    print()
